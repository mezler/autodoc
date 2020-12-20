# -*- coding: utf-8 -*- 
import shutil # to save it locally
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from collections import OrderedDict
import csv
import time
import sys

links = ["https://panzishop.hu/index.php?page=shopcategory&id=01&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=02&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=03&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=03&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=04&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=05&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=06&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=07&g=",
        "https://panzishop.hu/index.php?page=shopcategory&id=09&g="]

#  -------- FUNCTIONS --------

def check_exists_by_xpath(driver, xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

#  -------- / FUNCTIONS -------

book = Workbook()
sheet = book.active
rows = []

options = Options()
options.headless = True
driver = webdriver.Chrome('chromedriver', chrome_options=options) # Optional argument, if not specified will search path.
driver2 = webdriver.Chrome('chromedriver', chrome_options=options) # Optional argument, if not specified will search path.
dict = {}
arr = []

# BEJELENTKEZÉS -- DRIVER
driver.get( "https://panzishop.hu/" )
username = driver.find_element_by_id("system_login")
username.clear()
username.send_keys("gigor.lorant@gmail.com")
password = driver.find_element_by_id("system_pass")
password.clear()
password.send_keys("wokxy3-Hejfom-fezdiw")
driver.find_element_by_id("shopuser_login_button").click()

# BEJELENTKEZÉS -- DRIVER2
driver2.get( "https://panzishop.hu/" )
username = driver2.find_element_by_id("system_login")
username.clear()
username.send_keys("gigor.lorant@gmail.com")
password = driver2.find_element_by_id("system_pass")
password.clear()
password.send_keys("wokxy3-Hejfom-fezdiw")
driver2.find_element_by_id("shopuser_login_button").click()



book = Workbook()
sheet = book.active

rows = []

links2 = []
for url in links:
    driver.get( url )
    alist = driver.find_elements_by_xpath("//div[starts-with(@class, 'p-item-img')]/a")
    for q in alist:
        links2.append( q.get_attribute("href") )
print ( "1. level done...")

links3 = []
for url in links2:
    driver.get( url )
    alist = driver.find_elements_by_xpath("//div[starts-with(@class, 'p-item-img')]/a")
    if len( alist ) == 0:
        links3.append( url )
    for q in alist:
        links3.append( q.get_attribute("href") )
print ( "2. level done...")

links4 = []
for url in links3:
    driver.get( url )
    alist = driver.find_elements_by_xpath("//div[starts-with(@class, 'page-number')]")
    if len(alist) == 0:
        links4.append( url )
    bol = True
    counter = 0
    if len(alist) > 0:    
        alist1 = alist[0].find_elements_by_xpath(".//ul/li")
        baseURL = alist1[0].find_element_by_xpath(".//a" ).get_attribute ('href')
        while bol:
            driver2.get ( baseURL.rstrip( "0" ) + str( counter ) )
            if check_exists_by_xpath (driver2, "//div[starts-with(@class, 'shop-item-container')]") == False:
                bol = False
            if bol:  
                links4.append ( baseURL.rstrip( "0" ) + str( counter ) ) 
                # print ( baseURL.rstrip( "0" ) + str( counter ) ) 
            counter += 1

counter = 0
links5 = []
name = ""
category = ""
price = ""
secd = ""
description = ""

with open('productlinks2.txt', 'a') as file:
    for k in links4:
        counter += 1
        print ( counter )
        driver2.get( k )
        prodLinks = driver2.find_elements_by_xpath("//a[starts-with(@class, 'button-green upp')]")
        for p in prodLinks:
            # print ( p.get_attribute('href') )
            # links5.append ( str(counter) + " : " +  p.get_attribute('href') )
            # driver.get( p.get_attribute('href') )
            file.write( p.get_attribute('href') + '\n')
            # if len( driver.find_elements_by_xpath("//a") ) < 3:
            #     continue
            # # ár
            # prices = driver.find_elements_by_xpath("//div[starts-with(@class, 'price')]/span")
            # price = prices[0].text
            # # név
            # name = driver.find_element_by_xpath("//div[starts-with(@class, 'col-md-5')]/H1").text
            # print ( str (counter) + " :  " + name + " -- " + price)
            # # kategória
            # cats = driver.find_elements_by_xpath("//div[starts-with(@class, 'page-title-address text-right')]/a")
            # category = ""
            # for cat in cats:
            #     category = category + "*" + cat.text
            # # leírás
            # descs = driver.find_elements_by_xpath("//div[starts-with(@class, 'col-md-5')]/p")
            # for d in descs:
            #     description = description + " " + d.text
            # # cikkszám
            # sku = driver.find_element_by_xpath("//span[starts-with(@class, 'gray-italic')]").get_attribute('innerHTML')
            
            # print ( sku )
            # print ( name )
            # print ( description )
            # print ( price )
            # print ( category )

            # arr.append ( [ sku, name, description, price,  category ] )
#             rows.insert( 0, ( sku, name, description, price,  category ))

# for row in rows:
#     sheet.append(row)

# book.save('panzi2.xlsx')


    # print ( k )
    # driver2.get( k )
    # name = driver2.get( k )
    # prices = driver.find_elements_by_xpath("//div[starts-with(@class, 'price')]/span")
    # price = prices[0].text
    # print ( name + " -- " + price)
           
        


        # for q in alist1:
        #     print ( q.find_element_by_xpath(".//a" ).get_attribute ('href') )
        #     links4.append( q.find_element_by_xpath(".//a" ).get_attribute ('href') )

        # if len(alist1) == 8:
        #     hiddenURL1 = alist1[6].find_element_by_xpath(".//a").get_attribute ( "href" )
        #     driver2.get( hiddenURL1 )
        #     alist_hiddenurl1 = driver2.find_elements_by_xpath("//div[starts-with(@class, 'page-number')]")
        #     l_hidden_1 = alist_hiddenurl1[0].find_elements_by_xpath(".//ul/li")

            # if len( l_hidden_1 ) == 8:
            #     for q in l_hidden_1:
            #         print ( q.find_element_by_xpath(".//a" ).get_attribute ('href') )
            #     links4.append( q.find_element_by_xpath(".//a" ).get_attribute ('href') ) UA MINT 73

            # if len( l_hidden_1 ) > 8:
            #     for q in l_hidden_1:
            #         links4.append( q.find_element_by_xpath(".//a" ).get_attribute ('href') )
            #     pagerNO = len( l_hidden_1 ) 
            #     hiddenURL2 = l_hidden_1[ pagerNO - 2 ].find_element_by_xpath(".//a").get_attribute ( "href" )
            #     print ( "**************")
            #     print( "hidden url  1 : " + hiddenURL1 )
            #     print( "hidden url  2 : " + hiddenURL2 )
            #     print ( "**************")


        # if len(alist1) < 8:       
        #     for q in alist1:
        #         print ( q.find_element_by_xpath(".//a" ).get_attribute ('href') )
        #         links4.append( q.find_element_by_xpath(".//a" ).get_attribute ('href') )

# print ( "*****************************************************")
# for i in list(OrderedDict.fromkeys(links4)):
#     print ( i )


exit()


# productLinkArray = []
# driver.get("https://www.regiojatek.hu/kategoriak.html")
# list = driver.find_elements_by_xpath("//div[starts-with(@id, 'category_')]")

# print ( "1. szakasz elkezdődött...")
# for item in list:
#     for u in item.find_elements_by_xpath(".//a[starts-with(@id, 'subitem_')]"):
#         arr.append( u.get_attribute("href") )

# print ( "2. szakasz elkezdődött...")
# for item in arr:   
#     print ( "2.1. szakasz..." + item )
#     driver.get(item)
#     for z in driver.find_elements_by_class_name("product-box"):
#         print ( "2.2. szakasz..." )
#         ls = z.find_elements_by_xpath(".//a")
#         productLinkArray.append(ls[0].get_attribute( "href") )
#         # print ( ls[0].get_attribute( "href"))
#     for p in driver.find_elements_by_xpath("//ul[(@class='pagination')]/li[not(@class='active')]/a"):
#         print ( "2.3. szakasz..." )
#         if len(p.find_elements_by_xpath(".//span")) == 0:
#             driver2.get( p.get_attribute( "href") )
#             for ww in driver2.find_elements_by_class_name("product-box"):
#                 ls2 = ww.find_elements_by_xpath(".//a")
#                 productLinkArray.append(ls2[0].get_attribute( "href") )
#                 # print (  ls2[0].get_attribute( "href") )

# print ( "3. szakasz elkezdődött...", len(productLinkArray))

# count = 0
# description = ""
# cikkszam = ""
# category = ""
# str1 = ""
# name = ""
# noOfImgs = 0
# lineArr = []

# for item in productLinkArray:
#     count = count + 1
#     print count
#     sys.stdout.flush()
#     # if count == 200:
#     #     break
#     print ( str(count) )
#     sys.stdout.flush()
#     print ( "Item: ", item)
#     sys.stdout.flush()
#     driver.get( item )

#     # GET CIKKSZAM
#     if check_exists_by_xpath(driver, "//dl[starts-with(@class, 'product-data')]"):
#         elem = driver.find_element_by_xpath("//dl[starts-with(@class, 'product-data')]")
#         elem2 = elem.find_elements_by_xpath(".//dd")
#         cikkszam = elem2[0].find_element_by_xpath(".//span").text
#         str1 = cikkszam + "**" 
#         print ("cikkszam   --  ", cikkszam)
#     else:
#         file = open('skippedones.txt', 'a')
#         file.write(str1.encode('utf8'))
#         file.close() 
#         continue

#     # GET DESCRIPTION
#     desc = driver.find_elements_by_class_name("prodduct-description")
#     if len(desc) > 0:
#         description = desc[0].text
#     if len(desc) == 0:
#         description = ""
#     str1 = str1 + description + "**" 

#     # GET CATEGORY
#     cats = elem2[3].find_elements_by_xpath(".//a")
#     for cat in cats:
#         print ( "+++++ ", cat.text )
#         category = category + "|" + cat.text + "|"
#     str1 = str1 + category + "**" 

#     # GET NAME
#     name_element = driver.find_element_by_xpath("//div[starts-with(@class, 'col-md-7 col-sm-6')]/div/div/div/h1/span")
#     name = name_element.text
#     str1 = str1 + name + "**" 
#     print ("NAME NAME NAME : ", name)

#     # GET PRICE
#     price_element = driver.find_element_by_xpath("//div[starts-with(@class, 'product-prices product-price')]/span")
#     price = price_element.text
#     str1 = str1 + price + "**" 
#     print ("NAME NAME NAME : ", name)

#     # GET NUM OF IMGS
#     imgs = driver.find_elements_by_xpath("//div[starts-with(@class, 'carousel-inner carousel-inner-gallery')]/div")
#     noOfImgs = len ( imgs )
#     str1 = str1 + str(noOfImgs) + "**" 
#     print (" image darabszám : ", str(noOfImgs))

#     # GET STOCK INFO
#     stock = driver.find_element_by_xpath("//div[starts-with(@class, 'product-stock-info')]/h4")
#     str1 = str1 + stock.get_attribute('innerHTML')
#     print ("Stock info: ", stock.get_attribute('innerHTML'))

#     lineArr.append ( [cikkszam, price, description, category, name, noOfImgs, stock.get_attribute('innerHTML')])

#     # file_object = open('file_3.txt', 'a')
#     # file_object.write(str1.encode('utf8') + "\n")
#     # file_object.close()   

#     rows.insert( 0, (cikkszam, price, description, category, name, noOfImgs, stock.get_attribute('innerHTML')))
#     category = ""

# for row in rows:
#     sheet.append(row)

# book.save('appending_3.xlsx')



    


        
    
         






    
    # for item2 in driver.find_elements_by_class_name("product-data clearfix"):
    #     print ("+++ ", item2[0].get_attribute("innerHTML"), flush=True)



# VÉGRE !!!!!



# driver.close()