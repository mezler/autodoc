# -*- coding: utf-8 -*- 
import shutil # to save it locally
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from collections import OrderedDict
import csv
import time
import sys
import urllib.request
import requests

#  -------- FUNCTIONS --------

def check_exists_by_xpath(driver, xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

#  -------- / FUNCTIONS -------

book = Workbook()
sheet = book.active
rows = []

options = Options()
options.headless = True
driver = webdriver.Chrome('chromedriver', chrome_options=options) # Optional argument, if not specified will search path.
driver2 = webdriver.Chrome('chromedriver', chrome_options=options) # Optional argument, if not specified will search path.
dict = {}
arr = []

# BEJELENTKEZÉS -- DRIVER
driver.get( "https://panzishop.hu/" )
username = driver.find_element_by_id("system_login")
username.clear()
username.send_keys("gigor.lorant@gmail.com")
password = driver.find_element_by_id("system_pass")
password.clear()
password.send_keys("wokxy3-Hejfom-fezdiw")
driver.find_element_by_id("shopuser_login_button").click()

book = Workbook()
sheet = book.active
rows = []

counter = 0
links5 = []
name = ""
category = ""
price = ""
secd = ""
description = ""

file = open('productlinks3.txt', 'r')
Lines = file.readlines() 

for p in Lines:
    counter += 1
    # if ( counter > 1000):
    #     break
    print ( counter )
    driver.get( p )
    if len( driver.find_elements_by_xpath("//a") ) < 3:
        continue
 
    sku = driver.find_element_by_xpath("//span[starts-with(@class, 'gray-italic')]/strong").text
    # e = driver.find_element_by_xpath("//div[contains(@class, 'col-md-4') and not(@class='col-sm-2') and not(@class='col-sm-5' and not(@class='col-sm-6'))]")
    img = driver.find_element_by_xpath("//div[starts-with(@class, 'shop shop-single')]/div/div/div/a/img").get_attribute('src')
    # img = driver.find_element_by_class_name('col-md-4')
    print ( img )

    # print ( "https://panzishop.hu/" + img )

    try:
        response = requests.get(  img )
    except requests.exceptions.RequestException:
        continue

    file = open(sku + ".jpg", "wb")
    file.write(response.content)
    file.close()

    # response = requests.get(  img )

   




    # response = requests.get("https://i.imgur.com/ExdKOOz.png")

    # file = open("sample_image.png", "wb")
    # file.write(response.content)
    # file.close()

    # response = urllib.request.urlretrieve(img, "temp.jpg")

    # file = open(sku + ".jpg", "wb")
    # file.write(response.content)
    # file.close()
    # urllib.request.urlretrieve(img, sku + ".jpg")

    print ( sku )

